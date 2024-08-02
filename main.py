import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection

def compare_sheets(file_path):
    """
    Compares all sheets in the given Excel workbook and highlights differences.
    Differences are marked in a new sheet named 'Differences'.
    
    Parameters:
    file_path (str): The path to the Excel workbook to compare.
    """
    # Load the workbook and get sheet names
    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames

    # Create or get the 'Differences' sheet
    if 'Differences' in sheet_names:
        differences_sheet = wb['Differences']
    else:
        differences_sheet = wb.create_sheet(title='Differences')

    # Fill color for highlighting differences (Orange)
    fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    def copy_format(src_cell, dest_cell):
        """
        Copies the format from source cell to destination cell, excluding the fill.
        
        Parameters:
        src_cell (Cell): The source cell from which format is copied.
        dest_cell (Cell): The destination cell to which format is applied.
        """
        if src_cell.has_style:
            # Copy font style
            dest_cell.font = Font(
                name=src_cell.font.name,
                size=src_cell.font.size,
                bold=src_cell.font.bold,
                italic=src_cell.font.italic,
                vertAlign=src_cell.font.vertAlign,
                underline=src_cell.font.underline,
                strike=src_cell.font.strike,
                color=src_cell.font.color
            )

            # Copy border style
            def get_side(side):
                return Side(style=side.style, color=side.color) if side else None

            dest_cell.border = Border(
                left=get_side(src_cell.border.left),
                right=get_side(src_cell.border.right),
                top=get_side(src_cell.border.top),
                bottom=get_side(src_cell.border.bottom),
                diagonal=get_side(src_cell.border.diagonal),
                diagonal_direction=src_cell.border.diagonal_direction,
                outline=src_cell.border.outline,
                vertical=get_side(src_cell.border.vertical),
                horizontal=get_side(src_cell.border.horizontal)
            )

            # Set fill color to highlight differences
            dest_cell.fill = fill
            # Copy number format
            dest_cell.number_format = src_cell.number_format
            # Copy protection style
            dest_cell.protection = Protection(
                locked=src_cell.protection.locked,
                hidden=src_cell.protection.hidden
            )
            # Copy alignment style
            dest_cell.alignment = Alignment(
                horizontal=src_cell.alignment.horizontal,
                vertical=src_cell.alignment.vertical,
                text_rotation=src_cell.alignment.text_rotation,
                wrap_text=src_cell.alignment.wrap_text,
                shrink_to_fit=src_cell.alignment.shrink_to_fit,
                indent=src_cell.alignment.indent
            )

    # Copy column headings to the 'Differences' sheet
    first_sheet = wb[sheet_names[0]]
    for col in range(1, first_sheet.max_column + 1):
        header_cell = first_sheet.cell(row=1, column=col)
        diff_header_cell = differences_sheet.cell(row=1, column=col, value=header_cell.value)
        copy_format(header_cell, diff_header_cell)
        diff_header_cell.fill = fill  # Highlight headers as well

    # Compare sheets
    for sheet_name in sheet_names:
        if sheet_name == 'Differences':
            continue

        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column

        # Compare each cell in the sheet
        for row in range(2, max_row + 1):  # Skip header row
            for col in range(1, max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value

                for other_sheet_name in sheet_names:
                    if other_sheet_name == sheet_name or other_sheet_name == 'Differences':
                        continue

                    other_sheet = wb[other_sheet_name]
                    if row <= other_sheet.max_row and col <= other_sheet.max_column:
                        other_cell = other_sheet.cell(row=row, column=col)
                        if cell_value != other_cell.value:
                            # Mark differences in the 'Differences' sheet
                            diff_cell_1 = differences_sheet.cell(row=row, column=col, value=cell_value)
                            diff_cell_2 = differences_sheet.cell(row=row, column=col + 1, value=other_cell.value)

                            # Set fill color to highlight differences
                            diff_cell_1.fill = fill
                            diff_cell_2.fill = fill

                            # Copy formats
                            copy_format(cell, diff_cell_1)
                            copy_format(other_cell, diff_cell_2)

    # Save the workbook with differences
    wb.save(file_path)
    print(f"Differences highlighted in orange and saved in {file_path}")

# Example usage
if __name__ == "__main__":
    file_path = '<Add your Workbook Name>.xlsx'
    compare_sheets(file_path)
